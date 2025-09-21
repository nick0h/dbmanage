from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import FormView, ListView, UpdateView, DetailView, DeleteView, TemplateView
from django.urls import reverse_lazy
from .forms import RequestForm, RequestEditForm, RequestSearchForm, StudyEditForm, AntibodyForm, RequestorForm, TissueForm, StatusForm, AssigneeForm, ProbeForm, PriorityForm, SectioningRequestSearchForm, EmbeddingRequestSearchForm, EmbeddingRequestForm, SectioningRequestForm, StainingRequestSearchForm, EmbeddingRequestEditForm, SectioningRequestEditForm
from .models import Request, Status, Study, Requestor, Antibody, Tissue, Assignee, Probe, Priority, StainingRequest, EmbeddingRequest, SectioningRequest, StainingRequestChangeLog, EmbeddingRequestChangeLog, SectioningRequestChangeLog
from django.http import JsonResponse
from django.db.models import Q
from datetime import datetime
import re
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import openpyxl
import os

# Create your views here.

def home(request):
    return render(request, 'requests_app/home.html')



def data_management(request):
    return render(request, 'requests_app/data_management.html')

# Request Views
class RequestListView(ListView):
    model = Request
    template_name = 'requests_app/request_list.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = Request.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', '-created_at')
        order = self.request.GET.get('order', 'desc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'created_at', 'requestor__name', 'description', 'antibody__name', 'probe__name', 'study__title', 'tissue__name', 'status__status', 'priority__value', 'assigned_to__name']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = '-created_at'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        context['current_order'] = self.request.GET.get('order', 'desc')
        return context

class RequestDetailView(DetailView):
    model = Request
    template_name = 'requests_app/request_detail.html'
    context_object_name = 'request'

class RequestCreateView(FormView):
    template_name = 'requests_app/request_form.html'
    form_class = RequestForm
    success_url = reverse_lazy('request_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['today_date'] = datetime.now().date()
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['max_tissues'] = 10
        context['assignees'] = Assignee.objects.all().order_by('name')
        context['probes'] = Probe.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        request_obj = form.save(commit=False)
        request_obj.status = Status.get_default_status()
        
        # Collect additional tissues from form data
        additional_tissues = []
        for key, value in self.request.POST.items():
            if key.startswith('tissue_') and value:  # tissue_0, tissue_1, etc.
                try:
                    tissue_id = int(value)
                    tissue = Tissue.objects.get(pk=tissue_id)
                    additional_tissues.append(tissue.name)
                except (ValueError, Tissue.DoesNotExist):
                    pass
        
        # Collect links from form data
        links = []
        for key, value in self.request.POST.items():
            if key.startswith('link_') and value:  # link_0, link_1, etc.
                links.append(value)
        
        # Populate the data JSONField with form data
        request_obj.data = {
            'date': self.request.POST.get('date'),
            'description': form.cleaned_data.get('description'),
            'special_request': form.cleaned_data.get('special_request'),
            'tissues': additional_tissues,
            'links': links,
        }
        
        # Store links in the links field
        if links:
            request_obj.links = links
        
        request_obj.save()
        return super().form_valid(form)

# Staining Request Create View
class StainingRequestCreateView(FormView):
    template_name = 'requests_app/request_form.html'
    form_class = RequestForm
    success_url = reverse_lazy('staining_requests')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['today_date'] = datetime.now().date()
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['max_tissues'] = 10
        context['assignees'] = Assignee.objects.all().order_by('name')
        context['probes'] = Probe.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        request_obj = form.save(commit=False)
        request_obj.status = Status.get_default_status()
        
        # Collect additional tissues from form data
        additional_tissues = []
        for key, value in self.request.POST.items():
            if key.startswith('tissue_') and value:  # tissue_0, tissue_1, etc.
                try:
                    tissue_id = int(value)
                    tissue = Tissue.objects.get(pk=tissue_id)
                    additional_tissues.append(tissue.name)
                except (ValueError, Tissue.DoesNotExist):
                    pass
        
        # Populate the data JSONField with form data
        request_obj.data = {
            'date': self.request.POST.get('date'),
            'description': form.cleaned_data.get('description'),
            'special_request': form.cleaned_data.get('special_request'),
            'tissues': additional_tissues,
        }
        
        request_obj.save()
        
        # Log the creation
        StainingRequestChangeLog.log_change(
            request=request_obj,
            change_type='created',
            description='Staining request created'
        )
        
        return super().form_valid(form)

# Embedding Request Create View
class EmbeddingRequestCreateView(FormView):
    template_name = 'requests_app/embedding_request_form.html'
    form_class = EmbeddingRequestForm
    success_url = reverse_lazy('embedding_requests')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tissues'] = Tissue.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        request_obj = form.save(commit=False)
        request_obj.save()
        
        # Handle tissues exactly like staining request
        # Primary tissue from form
        primary_tissue_id = self.request.POST.get('tissue')
        if primary_tissue_id:
            primary_tissue = Tissue.objects.get(pk=primary_tissue_id)
            request_obj.tissues.add(primary_tissue)
        
        # Additional tissues from JavaScript-created fields
        additional_tissues = []
        for key, value in self.request.POST.items():
            if key.startswith('tissue_') and value:  # tissue_0, tissue_1, etc.
                try:
                    tissue_id = int(value)
                    tissue = Tissue.objects.get(pk=tissue_id)
                    additional_tissues.append(tissue)
                except (ValueError, Tissue.DoesNotExist):
                    pass
        
        # Add additional tissues
        if additional_tissues:
            request_obj.tissues.add(*additional_tissues)
        
        # Handle links
        links = []
        for key, value in self.request.POST.items():
            if key.startswith('link_') and value:  # link_0, link_1, etc.
                links.append(value)
        
        if links:
            request_obj.links = links
            request_obj.save()
        
        # Log the creation
        EmbeddingRequestChangeLog.log_change(
            request=request_obj,
            change_type='created',
            description='Embedding request created'
        )
        
        return super().form_valid(form)

# Sectioning Request Create View
class SectioningRequestCreateView(FormView):
    template_name = 'requests_app/sectioning_request_form.html'
    form_class = SectioningRequestForm
    success_url = reverse_lazy('sectioning_requests')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['tissues'] = Tissue.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        request_obj = form.save(commit=False)
        request_obj.save()
        
        # Handle tissues exactly like staining request
        # Primary tissue from form
        primary_tissue_id = self.request.POST.get('tissue')
        if primary_tissue_id:
            primary_tissue = Tissue.objects.get(pk=primary_tissue_id)
            request_obj.tissues.add(primary_tissue)
        
        # Additional tissues from JavaScript-created fields
        additional_tissues = []
        for key, value in self.request.POST.items():
            if key.startswith('tissue_') and value:  # tissue_0, tissue_1, etc.
                try:
                    tissue_id = int(value)
                    tissue = Tissue.objects.get(pk=tissue_id)
                    additional_tissues.append(tissue)
                except (ValueError, Tissue.DoesNotExist):
                    pass
        
        # Add additional tissues
        if additional_tissues:
            request_obj.tissues.add(*additional_tissues)
        
        # Handle links
        links = []
        for key, value in self.request.POST.items():
            if key.startswith('link_') and value:  # link_0, link_1, etc.
                links.append(value)
        
        if links:
            request_obj.links = links
            request_obj.save()
        
        # Log the creation
        SectioningRequestChangeLog.log_change(
            request=request_obj,
            change_type='created',
            description='Sectioning request created'
        )
        
        return super().form_valid(form)

class RequestUpdateView(UpdateView):
    model = Request
    form_class = RequestEditForm
    template_name = 'requests_app/request_edit.html'
    success_url = reverse_lazy('request_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['request_obj'] = self.object
        context['antibodies'] = Antibody.objects.all().order_by('name')
        context['studies'] = Study.objects.all().order_by('study_id')
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['assignees'] = Assignee.objects.all().order_by('name')
        context['probes'] = Probe.objects.all().order_by('name')
        return context
    
    def form_valid(self, form):
        print("DEBUG: RequestUpdateView form_valid called")
        # Get the original object before saving
        original_obj = Request.objects.get(pk=self.object.pk)
        
        # Debug: print all POST data
        print(f"DEBUG: All POST keys: {list(self.request.POST.keys())}")
        
        # Handle links
        links = []
        for key, value in self.request.POST.items():
            if key.startswith('link_') and value:  # link_0, link_1, etc.
                links.append(value)
        
        # Debug: print what we're collecting
        print(f"DEBUG: Collected links: {links}")
        print(f"DEBUG: POST data with 'link': {[k for k in self.request.POST.keys() if 'link' in k]}")
        
        # Always update links (even if empty to handle removals)
        self.object.links = links
        self.object.save()
        print(f"DEBUG: Saved links to database: {self.object.links}")
        
        # Save the form
        response = super().form_valid(form)
        return response

class RequestDeleteView(DeleteView):
    model = Request
    template_name = 'requests_app/request_confirm_delete.html'
    success_url = reverse_lazy('request_list')
    context_object_name = 'request'

class RequestSearchView(ListView):
    model = Request
    template_name = 'requests_app/request_search.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Create form with GET data
        form = RequestSearchForm(self.request.GET)
        
        context['form'] = form
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['assignees'] = Assignee.objects.all().order_by('name')
        context['antibodies'] = Antibody.objects.all().order_by('name')
        context['probes'] = Probe.objects.all().order_by('name')
        context['studies'] = Study.objects.all().order_by('study_id')
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['priorities'] = Priority.objects.all().order_by('value')
        return context

    def get_queryset(self):
        queryset = Request.objects.all()
        form = RequestSearchForm(self.request.GET)
        
        if form.is_valid():
            request_id = form.cleaned_data.get('request_id')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            requestor = form.cleaned_data.get('requestor')
            description = form.cleaned_data.get('description')
            antibody = form.cleaned_data.get('antibody')
            probe = form.cleaned_data.get('probe')
            tissue = form.cleaned_data.get('tissue')
            study = form.cleaned_data.get('study')
            status = form.cleaned_data.get('status')
            priority = form.cleaned_data.get('priority')
            assigned_to = form.cleaned_data.get('assigned_to')
            special_request = form.cleaned_data.get('special_request')
            notes = form.cleaned_data.get('notes')

            # Apply filters
            if request_id:
                queryset = queryset.filter(key=request_id)
            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)
            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)
            if requestor:
                queryset = queryset.filter(requestor=requestor)
            if description:
                queryset = queryset.filter(description__icontains=description)
            if antibody:
                queryset = queryset.filter(antibody=antibody)
            if probe:
                queryset = queryset.filter(probe=probe)
            if tissue:
                queryset = queryset.filter(tissue=tissue)
            if study:
                queryset = queryset.filter(study=study)
            if status:
                queryset = queryset.filter(status=status)
            if priority:
                queryset = queryset.filter(priority=priority)
            if assigned_to:
                queryset = queryset.filter(assigned_to=assigned_to)
            if special_request:
                queryset = queryset.filter(special_request__icontains=special_request)
            if notes:
                queryset = queryset.filter(notes__icontains=notes)

        return queryset.order_by('-created_at')



class RequestHomeView(ListView):
    model = Request
    template_name = 'requests_app/request_home.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get sort parameters
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        context['current_order'] = self.request.GET.get('order', 'desc')
        
        return context

    def get_queryset(self):
        queryset = Request.objects.all()
        
        # Apply sorting
        sort_by = self.request.GET.get('sort', '-created_at')
        order = self.request.GET.get('order', 'desc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'created_at', 'requestor__name', 'description', 'antibody__name', 'probe__name', 'study__title', 'tissue__name', 'status__status', 'priority__value', 'assigned_to__name']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = '-created_at'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

class SimpleRequestHomeView(ListView):
    model = Request
    template_name = 'requests_app/simple_request_home.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get sort parameters
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        context['current_order'] = self.request.GET.get('order', 'desc')
        
        return context

    def get_queryset(self):
        queryset = Request.objects.all()
        
        # Apply sorting
        sort_by = self.request.GET.get('sort', '-created_at')
        order = self.request.GET.get('order', 'desc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'created_at', 'requestor__name', 'description', 'antibody__name', 'probe__name', 'study__title', 'tissue__name', 'status__status', 'priority__value', 'assigned_to__name']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = '-created_at'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

# Study Views
class StudyUpdateView(UpdateView):
    model = Study
    form_class = StudyEditForm
    template_name = 'requests_app/study_edit.html'
    success_url = reverse_lazy('study_list')

class StudyListView(ListView):
    model = Study
    template_name = 'requests_app/study_list.html'
    context_object_name = 'studies'

    def get_queryset(self):
        queryset = Study.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'key')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'study_id', 'title']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'key'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'key')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class StudyDeleteView(DeleteView):
    model = Study
    template_name = 'requests_app/study_confirm_delete.html'
    success_url = reverse_lazy('study_list')
    context_object_name = 'study'

def study_create(request):
    if request.method == 'POST':
        study_id = request.POST.get('study_id', '').strip()
        title = request.POST.get('title', '').strip()
        
        study_id = sanitize_input(study_id)
        title = sanitize_input(title)
        
        if validate_input(study_id) and validate_input(title):
            Study.objects.create(study_id=study_id, title=title)
            return redirect('study_list')
        else:
            return JsonResponse({'error': 'Invalid input'}, status=400)
    
    return render(request, 'requests_app/study_form.html', {'title': 'Add New Study'})

# Requestor Views
class RequestorListView(ListView):
    model = Requestor
    template_name = 'requests_app/requestor_list.html'
    context_object_name = 'requestors'

    def get_queryset(self):
        queryset = Requestor.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'key')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'name']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'key'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'key')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class RequestorUpdateView(UpdateView):
    model = Requestor
    form_class = RequestorForm
    template_name = 'requests_app/requestor_edit.html'
    success_url = reverse_lazy('requestor_list')

class RequestorDeleteView(DeleteView):
    model = Requestor
    template_name = 'requests_app/requestor_confirm_delete.html'
    success_url = reverse_lazy('requestor_list')
    context_object_name = 'requestor'

def sanitize_input(input_str):
    # Remove HTML tags
    clean = re.compile('<.*?>')
    return re.sub(clean, '', input_str)

def validate_input(input_str):
    # Check if input is not empty and contains only allowed characters
    if not input_str or len(input_str) > 256:
        return False
    # Allow letters, numbers, spaces, hyphens, and underscores
    return bool(re.match(r'^[a-zA-Z0-9\s\-_]+$', input_str))

def validate_import_input(input_str):
    # More permissive validation for imports - allows common characters in study titles and IDs
    if not input_str or len(input_str) > 256:
        return False
    # Allow letters, numbers, spaces, hyphens, underscores, parentheses, colons, periods, and commas
    return bool(re.match(r'^[a-zA-Z0-9\s\-_\(\):.,]+$', input_str))

def requestor_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        
        name = sanitize_input(name)
        
        if validate_input(name):
            Requestor.objects.create(name=name)
            return redirect('requestor_list')
        else:
            return JsonResponse({'error': 'Invalid input'}, status=400)
    
    return render(request, 'requests_app/requestor_form.html', {'title': 'Add New Requestor'})

# Antibody Views
class AntibodyListView(ListView):
    model = Antibody
    template_name = 'requests_app/antibody_list.html'
    context_object_name = 'antibodies'

    def get_queryset(self):
        queryset = Antibody.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'key')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'name', 'description', 'antigen', 'species', 'recognizes', 'vendor']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'key'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'key')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class AntibodyUpdateView(UpdateView):
    model = Antibody
    form_class = AntibodyForm
    template_name = 'requests_app/antibody_edit.html'
    success_url = reverse_lazy('antibody_list')

class AntibodyDeleteView(DeleteView):
    model = Antibody
    template_name = 'requests_app/antibody_confirm_delete.html'
    success_url = reverse_lazy('antibody_list')
    context_object_name = 'antibody'

def antibody_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        antigen = request.POST.get('antigen', '').strip()
        species = request.POST.get('species', '').strip()
        recognizes = request.POST.get('recognizes', '').strip()
        vendor = request.POST.get('vendor', '').strip()
        
        name = sanitize_input(name)
        description = sanitize_input(description)
        antigen = sanitize_input(antigen)
        species = sanitize_input(species)
        recognizes = sanitize_input(recognizes)
        vendor = sanitize_input(vendor)
        
        if validate_input(name) and validate_input(description) and validate_input(antigen) and validate_input(species) and validate_input(recognizes) and validate_input(vendor):
            Antibody.objects.create(name=name, description=description, antigen=antigen, species=species, recognizes=recognizes, vendor=vendor)
            return redirect('antibody_list')
        else:
            return JsonResponse({'error': 'Invalid input'}, status=400)
    
    return render(request, 'requests_app/antibody_form.html', {'title': 'Add New Antibody'})

# Tissue Views
class TissueListView(ListView):
    model = Tissue
    template_name = 'requests_app/tissue_list.html'
    context_object_name = 'tissues'

    def get_queryset(self):
        queryset = Tissue.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'key')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'name']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'key'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'key')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class TissueUpdateView(UpdateView):
    model = Tissue
    form_class = TissueForm
    template_name = 'requests_app/tissue_edit.html'
    success_url = reverse_lazy('tissue_list')

class TissueDeleteView(DeleteView):
    model = Tissue
    template_name = 'requests_app/tissue_confirm_delete.html'
    success_url = reverse_lazy('tissue_list')
    context_object_name = 'tissue'

def tissue_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        
        name = sanitize_input(name)
        
        if validate_input(name):
            Tissue.objects.create(name=name)
            return redirect('tissue_list')
        else:
            return JsonResponse({'error': 'Invalid input'}, status=400)
    
    return render(request, 'requests_app/tissue_form.html', {'title': 'Add New Tissue'})

# Status Views
class StatusListView(ListView):
    model = Status
    template_name = 'requests_app/status_list.html'
    context_object_name = 'statuses'

    def get_queryset(self):
        queryset = Status.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'key')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'status']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'key'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'key')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class StatusUpdateView(UpdateView):
    model = Status
    form_class = StatusForm
    template_name = 'requests_app/status_edit.html'
    success_url = reverse_lazy('status_list')

class StatusDeleteView(DeleteView):
    model = Status
    template_name = 'requests_app/status_confirm_delete.html'
    success_url = reverse_lazy('status_list')
    context_object_name = 'status'

def status_create(request):
    if request.method == 'POST':
        status = request.POST.get('status', '').strip()
        
        status = sanitize_input(status)
        
        if validate_input(status):
            Status.objects.create(status=status)
            return redirect('status_list')
        else:
            return JsonResponse({'error': 'Invalid input'}, status=400)
    
    return render(request, 'requests_app/status_form.html', {'title': 'Add New Status'})

# Assignee Views
class AssigneeListView(ListView):
    model = Assignee
    template_name = 'requests_app/assignee_list.html'
    context_object_name = 'assignees'

    def get_queryset(self):
        queryset = Assignee.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'key')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'name', 'email', 'department']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'key'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'key')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class AssigneeUpdateView(UpdateView):
    model = Assignee
    form_class = AssigneeForm
    template_name = 'requests_app/assignee_edit.html'
    success_url = reverse_lazy('assignee_list')

class AssigneeDeleteView(DeleteView):
    model = Assignee
    template_name = 'requests_app/assignee_confirm_delete.html'
    success_url = reverse_lazy('assignee_list')
    context_object_name = 'assignee'

def assignee_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        department = request.POST.get('department', '').strip()
        
        name = sanitize_input(name)
        email = sanitize_input(email)
        department = sanitize_input(department)
        
        if validate_input(name):
            Assignee.objects.create(name=name, email=email, department=department)
            return redirect('assignee_list')
        else:
            return JsonResponse({'error': 'Invalid input'}, status=400)
    
    return render(request, 'requests_app/assignee_form.html', {'title': 'Add New Assignee'})

# Probe Views
class ProbeListView(ListView):
    model = Probe
    template_name = 'requests_app/probe_list.html'
    context_object_name = 'probes'

    def get_queryset(self):
        queryset = Probe.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'key')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'name', 'description', 'sequence', 'target_gene', 'vendor', 'catalog_number', 'platform', 'target_region', 'number_of_pairs']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'key'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'key')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class ProbeUpdateView(UpdateView):
    model = Probe
    form_class = ProbeForm
    template_name = 'requests_app/probe_edit.html'
    success_url = reverse_lazy('probe_list')

class ProbeDeleteView(DeleteView):
    model = Probe
    template_name = 'requests_app/probe_confirm_delete.html'
    success_url = reverse_lazy('probe_list')
    context_object_name = 'probe'

def probe_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        sequence = request.POST.get('sequence', '').strip()
        target_gene = request.POST.get('target_gene', '').strip()
        vendor = request.POST.get('vendor', '').strip()
        catalog_number = request.POST.get('catalog_number', '').strip()
        platform = request.POST.get('platform', '').strip()
        target_region = request.POST.get('target_region', '').strip()
        number_of_pairs = request.POST.get('number_of_pairs', '').strip()
        
        name = sanitize_input(name)
        description = sanitize_input(description)
        sequence = sanitize_input(sequence)
        target_gene = sanitize_input(target_gene)
        vendor = sanitize_input(vendor)
        catalog_number = sanitize_input(catalog_number)
        platform = sanitize_input(platform)
        target_region = sanitize_input(target_region)
        number_of_pairs = sanitize_input(number_of_pairs)
        
        if validate_input(name):
            # Convert number_of_pairs to integer if provided
            pairs_value = None
            if number_of_pairs:
                try:
                    pairs_value = int(number_of_pairs)
                except ValueError:
                    pairs_value = None
            
            Probe.objects.create(
                name=name,
                description=description,
                sequence=sequence,
                target_gene=target_gene,
                vendor=vendor,
                catalog_number=catalog_number,
                platform=platform,
                target_region=target_region,
                number_of_pairs=pairs_value
            )
            return redirect('probe_list')
        else:
            return JsonResponse({'error': 'Invalid input'}, status=400)
    
    return render(request, 'requests_app/probe_form.html', {'title': 'Add New Probe'})

class PriorityListView(ListView):
    model = Priority
    template_name = 'requests_app/priority_list.html'
    context_object_name = 'priorities'

    def get_queryset(self):
        queryset = Priority.objects.all()
        
        # Get sort parameters
        sort_by = self.request.GET.get('sort', 'value')
        order = self.request.GET.get('order', 'asc')
        
        # Validate sort field to prevent injection
        allowed_fields = ['key', 'value', 'label', 'description']
        if sort_by.lstrip('-') not in [field.lstrip('-') for field in allowed_fields]:
            sort_by = 'value'
        
        # Apply sorting
        if order == 'asc' and sort_by.startswith('-'):
            sort_by = sort_by[1:]
        elif order == 'desc' and not sort_by.startswith('-'):
            sort_by = '-' + sort_by
            
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'value')
        context['current_order'] = self.request.GET.get('order', 'asc')
        return context

class PriorityUpdateView(UpdateView):
    model = Priority
    form_class = PriorityForm
    template_name = 'requests_app/priority_edit.html'
    success_url = reverse_lazy('priority_list')

class PriorityDeleteView(DeleteView):
    model = Priority
    template_name = 'requests_app/priority_confirm_delete.html'
    success_url = reverse_lazy('priority_list')
    context_object_name = 'priority'

def priority_create(request):
    if request.method == 'POST':
        form = PriorityForm(request.POST)
        if form.is_valid():
            # Sanitize input
            priority = form.save(commit=False)
            priority.value = sanitize_input(str(priority.value))
            priority.label = sanitize_input(priority.label)
            priority.description = sanitize_input(priority.description)
            
            # Validate input
            if not validate_input(priority.value) or not validate_input(priority.label):
                messages.error(request, 'Invalid input detected. Please check your data.')
                return render(request, 'requests_app/priority_form.html', {'form': form})
            
            priority.save()
            messages.success(request, 'Priority created successfully!')
            return redirect('priority_list')
    else:
        form = PriorityForm()
    
    return render(request, 'requests_app/priority_form.html', {'form': form})

def custom_404(request, exception):
    return render(request, "404.html", status=404)

def custom_500(request):
    return render(request, "500.html", status=500)

def import_studies_from_file(request):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'No file was uploaded.')
            return redirect('study_list')
        
        uploaded_file = request.FILES['file']
        
        # Check file extension
        if not uploaded_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload an Excel file (.xlsx)')
            return redirect('study_list')
        
        try:
            # Read the Excel file
            workbook = openpyxl.load_workbook(uploaded_file)
            sheet = workbook.active
            
            # Check if the first row contains the expected headers
            first_row = [cell.value for cell in sheet[1]]
            if len(first_row) < 2:
                messages.error(request, 'File formatting is incorrect for study')
                return redirect('study_list')
            
            # Check for required headers (case insensitive)
            first_row_lower = [str(cell).lower().strip() if cell else '' for cell in first_row]
            study_id_index = None
            title_index = None
            
            for i, header in enumerate(first_row_lower):
                if 'study id' in header:
                    study_id_index = i
                elif 'title' in header:
                    title_index = i
            
            if study_id_index is None or title_index is None:
                messages.error(request, 'File formatting is incorrect for study')
                return redirect('study_list')
            
            # Process each row starting from the second row
            imported_count = 0
            for row_num in range(2, sheet.max_row + 1):
                study_id_cell = sheet.cell(row=row_num, column=study_id_index + 1)
                title_cell = sheet.cell(row=row_num, column=title_index + 1)
                
                study_id = study_id_cell.value
                title = title_cell.value
                
                # Skip empty rows
                if not study_id or not title:
                    continue
                
                # Clean and validate the data
                study_id = str(study_id).strip()
                title = str(title).strip()
                
                if validate_import_input(study_id) and validate_import_input(title):
                    # Check if study already exists
                    if not Study.objects.filter(study_id=study_id).exists():
                        Study.objects.create(study_id=study_id, title=title)
                        imported_count += 1
            
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} studies.')
            else:
                messages.warning(request, 'No new studies were imported.')
                
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('study_list')
    
    return redirect('study_list')

def import_antibodies_from_file(request):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'No file was uploaded.')
            return redirect('antibody_list')
        
        uploaded_file = request.FILES['file']
        
        # Check file extension
        if not uploaded_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload an Excel file (.xlsx)')
            return redirect('antibody_list')
        
        try:
            # Read the Excel file
            workbook = openpyxl.load_workbook(uploaded_file)
            sheet = workbook.active
            
            # Check if the first row contains the expected headers
            first_row = [cell.value for cell in sheet[1]]
            if len(first_row) < 6:
                messages.error(request, 'File formatting is incorrect for antibody')
                return redirect('antibody_list')
            
            # Check for required headers (case insensitive)
            first_row_lower = [str(cell).lower().strip() if cell else '' for cell in first_row]
            name_index = None
            description_index = None
            antigen_index = None
            species_index = None
            recognizes_index = None
            vendor_index = None
            
            for i, header in enumerate(first_row_lower):
                if 'name' in header:
                    name_index = i
                elif 'description' in header:
                    description_index = i
                elif 'antigen' in header:
                    antigen_index = i
                elif 'species' in header:
                    species_index = i
                elif 'recognizes' in header:
                    recognizes_index = i
                elif 'vendor' in header:
                    vendor_index = i
            
            if (name_index is None or description_index is None or antigen_index is None or 
                species_index is None or recognizes_index is None or vendor_index is None):
                messages.error(request, 'File formatting is incorrect for antibody')
                return redirect('antibody_list')
            
            # Process each row starting from the second row
            imported_count = 0
            for row_num in range(2, sheet.max_row + 1):
                name_cell = sheet.cell(row=row_num, column=name_index + 1)
                description_cell = sheet.cell(row=row_num, column=description_index + 1)
                antigen_cell = sheet.cell(row=row_num, column=antigen_index + 1)
                species_cell = sheet.cell(row=row_num, column=species_index + 1)
                recognizes_cell = sheet.cell(row=row_num, column=recognizes_index + 1)
                vendor_cell = sheet.cell(row=row_num, column=vendor_index + 1)
                
                name = name_cell.value
                description = description_cell.value
                antigen = antigen_cell.value
                species = species_cell.value
                recognizes = recognizes_cell.value
                vendor = vendor_cell.value
                
                # Skip empty rows (name is required)
                if not name:
                    continue
                
                # Clean and validate the data
                name = str(name).strip()
                description = str(description).strip() if description else ''
                antigen = str(antigen).strip() if antigen else ''
                species = str(species).strip() if species else ''
                recognizes = str(recognizes).strip() if recognizes else ''
                vendor = str(vendor).strip() if vendor else ''
                
                if validate_import_input(name):
                    # Check if antibody already exists
                    if not Antibody.objects.filter(name=name).exists():
                        Antibody.objects.create(
                            name=name,
                            description=description,
                            antigen=antigen,
                            species=species,
                            recognizes=recognizes,
                            vendor=vendor
                        )
                        imported_count += 1
            
            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} antibodies.')
            else:
                messages.warning(request, 'No new antibodies were imported.')
                
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('antibody_list')
    
    return redirect('antibody_list')

# Staining Request Views
class StainingRequestsView(ListView):
    model = StainingRequest
    template_name = 'requests_app/staining_requests.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = StainingRequest.objects.all()
        
        # Get filter parameters
        selected_statuses = self.request.GET.getlist('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        sort_by = self.request.GET.get('sort', '-created_at')
        
        # Filter by status (only if statuses are selected)
        if selected_statuses:
            # Convert status names to status objects
            status_objects = Status.objects.filter(status__in=selected_statuses)
            queryset = queryset.filter(status__in=status_objects)
        
        # Filter by date range
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        # Apply sorting
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['selected_statuses'] = self.request.GET.getlist('status')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        return context

# Embedding Request Views
class EmbeddingRequestsView(ListView):
    model = EmbeddingRequest
    template_name = 'requests_app/embedding_requests.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = EmbeddingRequest.objects.all()
        
        # Get filter parameters
        selected_statuses = self.request.GET.getlist('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        sort_by = self.request.GET.get('sort', '-created_at')
        
        # Filter by status (only if statuses are selected)
        if selected_statuses:
            # Convert status names to status objects
            status_objects = Status.objects.filter(status__in=selected_statuses)
            queryset = queryset.filter(status__in=status_objects)
        
        # Filter by date range
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        # Apply sorting
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['selected_statuses'] = self.request.GET.getlist('status')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        return context

class EmbeddingRequestSearchView(ListView):
    model = EmbeddingRequest
    template_name = 'requests_app/embedding_request_search.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = EmbeddingRequestSearchForm(self.request.GET)
        context['form'] = form
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['assignees'] = Assignee.objects.all().order_by('name')
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['priorities'] = Priority.objects.all().order_by('value')
        return context

    def get_queryset(self):
        queryset = EmbeddingRequest.objects.all()
        form = EmbeddingRequestSearchForm(self.request.GET)
        
        if form.is_valid():
            request_id = form.cleaned_data.get('request_id')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            requestor = form.cleaned_data.get('requestor')
            tissue = form.cleaned_data.get('tissue')
            study = form.cleaned_data.get('study')
            status = form.cleaned_data.get('status')
            assigned_to = form.cleaned_data.get('assigned_to')
            special_request = form.cleaned_data.get('special_request')
            number_of_animals = form.cleaned_data.get('number_of_animals')
            take_down_date = form.cleaned_data.get('take_down_date')
            currently_in = form.cleaned_data.get('currently_in')
            date_of_xylene_etoh_change = form.cleaned_data.get('date_of_xylene_etoh_change')
            length_of_time_in_etoh = form.cleaned_data.get('length_of_time_in_etoh')

            # Apply filters
            if request_id:
                queryset = queryset.filter(key=request_id)
            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)
            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)
            if requestor:
                queryset = queryset.filter(requestor=requestor)
            if tissue:
                queryset = queryset.filter(tissues=tissue)
            if study:
                queryset = queryset.filter(study=study)
            if status:
                queryset = queryset.filter(status=status)
            if assigned_to:
                queryset = queryset.filter(assigned_to=assigned_to)
            if special_request:
                queryset = queryset.filter(special_request__icontains=special_request)
            if number_of_animals:
                queryset = queryset.filter(number_of_animals=number_of_animals)
            if take_down_date:
                queryset = queryset.filter(take_down_date=take_down_date)
            if currently_in != '':
                queryset = queryset.filter(currently_in=currently_in)
            if date_of_xylene_etoh_change:
                queryset = queryset.filter(date_of_xylene_etoh_change=date_of_xylene_etoh_change)
            if length_of_time_in_etoh:
                queryset = queryset.filter(length_of_time_in_etoh__icontains=length_of_time_in_etoh)

        return queryset.order_by('-created_at')

# Sectioning Request Views
class SectioningRequestsView(ListView):
    model = SectioningRequest
    template_name = 'requests_app/sectioning_requests.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = SectioningRequest.objects.all()
        
        # Get filter parameters
        selected_statuses = self.request.GET.getlist('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        sort_by = self.request.GET.get('sort', '-created_at')
        
        # Filter by status (only if statuses are selected)
        if selected_statuses:
            # Convert status names to status objects
            status_objects = Status.objects.filter(status__in=selected_statuses)
            queryset = queryset.filter(status__in=status_objects)
        
        # Filter by date range
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        # Apply sorting
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['selected_statuses'] = self.request.GET.getlist('status')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        return context

class SectioningRequestSearchView(ListView):
    model = SectioningRequest
    template_name = 'requests_app/sectioning_request_search.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        form = SectioningRequestSearchForm(self.request.GET)
        context['form'] = form
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['assignees'] = Assignee.objects.all().order_by('name')
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['priorities'] = Priority.objects.all().order_by('value')
        return context

    def get_queryset(self):
        queryset = SectioningRequest.objects.all()
        form = SectioningRequestSearchForm(self.request.GET)
        
        if form.is_valid():
            request_id = form.cleaned_data.get('request_id')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
            requestor = form.cleaned_data.get('requestor')
            tissue = form.cleaned_data.get('tissue')
            study = form.cleaned_data.get('study')
            status = form.cleaned_data.get('status')
            assigned_to = form.cleaned_data.get('assigned_to')
            special_request = form.cleaned_data.get('special_request')
            cut_surface_down = form.cleaned_data.get('cut_surface_down')
            sections_per_slide = form.cleaned_data.get('sections_per_slide')
            slides_per_block = form.cleaned_data.get('slides_per_block')
            other = form.cleaned_data.get('other')
            for_what = form.cleaned_data.get('for_what')

            # Apply filters
            if request_id:
                queryset = queryset.filter(key=request_id)
            if date_from:
                queryset = queryset.filter(created_at__gte=date_from)
            if date_to:
                queryset = queryset.filter(created_at__lte=date_to)
            if requestor:
                queryset = queryset.filter(requestor=requestor)
            if tissue:
                queryset = queryset.filter(tissues=tissue)
            if study:
                queryset = queryset.filter(study=study)
            if status:
                queryset = queryset.filter(status=status)
            if assigned_to:
                queryset = queryset.filter(assigned_to=assigned_to)
            if special_request:
                queryset = queryset.filter(special_request__icontains=special_request)
            if cut_surface_down:
                if cut_surface_down == 'Yes':
                    queryset = queryset.filter(cut_surface_down=True)
                elif cut_surface_down == 'No':
                    queryset = queryset.filter(cut_surface_down=False)
            if sections_per_slide:
                queryset = queryset.filter(sections_per_slide=sections_per_slide)
            if slides_per_block:
                queryset = queryset.filter(slides_per_block=slides_per_block)
            if other:
                queryset = queryset.filter(other__icontains=other)
            if for_what:
                queryset = queryset.filter(for_what=for_what)

        return queryset.order_by('-created_at')

# Staining Request Search View
class StainingRequestSearchView(ListView):
    model = Request
    template_name = 'requests_app/request_search.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = Request.objects.all()
        
        # Get search parameters
        request_id = self.request.GET.get('request_id')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        requestor = self.request.GET.get('requestor')
        description = self.request.GET.get('description')
        antibody = self.request.GET.get('antibody')
        probe = self.request.GET.get('probe')
        tissue = self.request.GET.get('tissue')
        study = self.request.GET.get('study')
        status = self.request.GET.get('status')
        priority = self.request.GET.get('priority')
        assigned_to = self.request.GET.get('assigned_to')
        special_request = self.request.GET.get('special_request')
        notes = self.request.GET.get('notes')
        
        # Apply filters
        if request_id:
            queryset = queryset.filter(key=request_id)
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        if requestor:
            queryset = queryset.filter(requestor=requestor)
        if description:
            queryset = queryset.filter(description__icontains=description)
        if antibody:
            queryset = queryset.filter(antibody=antibody)
        if probe:
            queryset = queryset.filter(probe=probe)
        if tissue:
            queryset = queryset.filter(tissue=tissue)
        if study:
            queryset = queryset.filter(study=study)
        if status:
            queryset = queryset.filter(status=status)
        if priority:
            queryset = queryset.filter(priority=priority)
        if assigned_to:
            queryset = queryset.filter(assigned_to=assigned_to)
        if special_request:
            queryset = queryset.filter(special_request__icontains=special_request)
        if notes:
            queryset = queryset.filter(notes__icontains=notes)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = StainingRequestSearchForm(self.request.GET)
        context['search_performed'] = any(self.request.GET.get(field) for field in ['request_id', 'date_from', 'date_to', 'requestor', 'description', 'antibody', 'probe', 'tissue', 'study', 'status', 'priority', 'assigned_to', 'special_request', 'notes'])
        return context
# Staining Request Detail, Edit, Delete Views
class StainingRequestDetailView(DetailView):
    model = Request
    template_name = 'requests_app/request_detail.html'
    context_object_name = 'request'

class StainingRequestEditView(UpdateView):
    model = Request
    form_class = RequestEditForm
    template_name = 'requests_app/request_edit.html'
    success_url = reverse_lazy('staining_requests')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['request_obj'] = self.object
        context['antibodies'] = Antibody.objects.all().order_by('name')
        context['studies'] = Study.objects.all().order_by('study_id')
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['assignees'] = Assignee.objects.all().order_by('name')
        context['probes'] = Probe.objects.all().order_by('name')
        context['priorities'] = Priority.objects.all().order_by('value')
        return context
    
    def form_valid(self, form):
        print("DEBUG: StainingRequestEditView form_valid called")
        # Get the original object before saving
        original_obj = Request.objects.get(pk=self.object.pk)
        
        # Debug: print all POST data
        print(f"DEBUG: All POST keys: {list(self.request.POST.keys())}")
        
        # Handle links
        links = []
        for key, value in self.request.POST.items():
            if key.startswith('link_') and value:  # link_0, link_1, etc.
                links.append(value)
        
        # Debug: print what we're collecting
        print(f"DEBUG: Collected links: {links}")
        print(f"DEBUG: POST data with 'link': {[k for k in self.request.POST.keys() if 'link' in k]}")
        
        # Always update links (even if empty to handle removals)
        self.object.links = links
        self.object.save()
        print(f"DEBUG: Saved links to database: {self.object.links}")
        
        # Save the form
        response = super().form_valid(form)
        
        # Determine what fields changed
        changed_fields = []
        for field in form.changed_data:
            if hasattr(original_obj, field) and hasattr(self.object, field):
                old_value = getattr(original_obj, field)
                new_value = getattr(self.object, field)
                if old_value != new_value:
                    changed_fields.append(field)
        
        # Log the change (always log, even if no fields detected as changed)
        StainingRequestChangeLog.log_change(
            request=self.object,
            change_type='updated',
            changed_fields=changed_fields,
            description=f'Updated fields: {", ".join(changed_fields) if changed_fields else "No specific fields detected as changed"}'
        )
        
        return response

class StainingRequestDeleteView(DeleteView):
    model = Request
    template_name = 'requests_app/request_confirm_delete.html'
    success_url = reverse_lazy('staining_requests')

# Embedding Request Detail, Edit, Delete Views
class EmbeddingRequestDetailView(DetailView):
    model = EmbeddingRequest
    template_name = 'requests_app/embedding_request_detail.html'
    context_object_name = 'request'

class EmbeddingRequestEditView(UpdateView):
    model = EmbeddingRequest
    form_class = EmbeddingRequestEditForm
    template_name = 'requests_app/embedding_request_edit.html'
    success_url = reverse_lazy('embedding_requests')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['request_obj'] = self.object
        context['requestors'] = Requestor.objects.all().order_by('name')
        context['studies'] = Study.objects.all().order_by('study_id')
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['assignees'] = Assignee.objects.all().order_by('name')
        return context
    
    def form_valid(self, form):
        print("DEBUG: EmbeddingRequestEditView form_valid called")
        # Get the original object before saving
        original_obj = EmbeddingRequest.objects.get(pk=self.object.pk)
        
        # Debug: print all POST data
        print(f"DEBUG: All POST keys: {list(self.request.POST.keys())}")
        
        # Handle links
        links = []
        for key, value in self.request.POST.items():
            if key.startswith('link_') and value:  # link_0, link_1, etc.
                links.append(value)
        
        # Debug: print what we're collecting
        print(f"DEBUG: Collected links: {links}")
        print(f"DEBUG: POST data with 'link': {[k for k in self.request.POST.keys() if 'link' in k]}")
        
        # Always update links (even if empty to handle removals)
        self.object.links = links
        self.object.save()
        print(f"DEBUG: Saved links to database: {self.object.links}")
        
        # Save the form
        response = super().form_valid(form)
        
        # Determine what fields changed
        changed_fields = []
        for field in form.changed_data:
            if hasattr(original_obj, field) and hasattr(self.object, field):
                old_value = getattr(original_obj, field)
                new_value = getattr(self.object, field)
                if old_value != new_value:
                    changed_fields.append(field)
        
        # Log the change (always log, even if no fields detected as changed)
        EmbeddingRequestChangeLog.log_change(
            request=self.object,
            change_type='updated',
            changed_fields=changed_fields,
            description=f'Updated fields: {", ".join(changed_fields) if changed_fields else "No specific fields detected as changed"}'
        )
        
        return response

class EmbeddingRequestDeleteView(DeleteView):
    model = EmbeddingRequest
    template_name = 'requests_app/request_confirm_delete.html'
    success_url = reverse_lazy('embedding_requests')

# Sectioning Request Detail, Edit, Delete Views
class SectioningRequestDetailView(DetailView):
    model = SectioningRequest
    template_name = 'requests_app/sectioning_request_detail.html'
    context_object_name = 'request'

class SectioningRequestEditView(UpdateView):
    model = SectioningRequest
    form_class = SectioningRequestEditForm
    template_name = 'requests_app/sectioning_request_edit.html'
    success_url = reverse_lazy('sectioning_requests')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['request_obj'] = self.object
        context['requestors'] = Requestor.objects.all().order_by('name')
        context['studies'] = Study.objects.all().order_by('study_id')
        context['tissues'] = Tissue.objects.all().order_by('name')
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['assignees'] = Assignee.objects.all().order_by('name')
        return context
    
    def form_valid(self, form):
        print("DEBUG: SectioningRequestEditView form_valid called")
        # Get the original object before saving
        original_obj = SectioningRequest.objects.get(pk=self.object.pk)
        
        # Debug: print all POST data
        print(f"DEBUG: All POST keys: {list(self.request.POST.keys())}")
        
        # Handle links
        links = []
        for key, value in self.request.POST.items():
            if key.startswith('link_') and value:  # link_0, link_1, etc.
                links.append(value)
        
        # Debug: print what we're collecting
        print(f"DEBUG: Collected links: {links}")
        print(f"DEBUG: POST data with 'link': {[k for k in self.request.POST.keys() if 'link' in k]}")
        
        # Always update links (even if empty to handle removals)
        self.object.links = links
        self.object.save()
        print(f"DEBUG: Saved links to database: {self.object.links}")
        
        # Save the form
        response = super().form_valid(form)
        
        # Determine what fields changed
        changed_fields = []
        for field in form.changed_data:
            if hasattr(original_obj, field) and hasattr(self.object, field):
                old_value = getattr(original_obj, field)
                new_value = getattr(self.object, field)
                if old_value != new_value:
                    changed_fields.append(field)
        
        # Log the change (always log, even if no fields detected as changed)
        SectioningRequestChangeLog.log_change(
            request=self.object,
            change_type='updated',
            changed_fields=changed_fields,
            description=f'Updated fields: {", ".join(changed_fields) if changed_fields else "No specific fields detected as changed"}'
        )
        
        return response

class SectioningRequestDeleteView(DeleteView):
    model = SectioningRequest
    template_name = 'requests_app/request_confirm_delete.html'
    success_url = reverse_lazy('sectioning_requests')


# Current Requests Views (Filtered - Exclude Complete)
class StainingRequestsCurrentView(ListView):
    model = StainingRequest
    template_name = 'requests_app/staining_requests.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = StainingRequest.objects.all()
        selected_statuses = self.request.GET.getlist('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        sort_by = self.request.GET.get('sort', '-created_at')
        
        # Filter by status
        if selected_statuses:
            # If statuses are selected, use only those (including Complete if selected)
            status_objects = Status.objects.filter(status__in=selected_statuses)
            queryset = queryset.filter(status__in=status_objects)
        else:
            # Default: exclude complete status when no statuses are selected
            complete_status = Status.objects.filter(status='Complete').first()
            if complete_status:
                queryset = queryset.exclude(status=complete_status)
        
        if date_from:
            queryset = queryset.filter(status_timestamp__gte=date_from)
        if date_to:
            queryset = queryset.filter(status_timestamp__lte=date_to)
        
        # Apply sorting
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['selected_statuses'] = self.request.GET.getlist('status')
        context['date_from'] = self.request.GET.get('date_from')
        context['date_to'] = self.request.GET.get('date_to')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        context['is_current_view'] = True
        return context


class EmbeddingRequestsCurrentView(ListView):
    model = EmbeddingRequest
    template_name = 'requests_app/embedding_requests.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = EmbeddingRequest.objects.all()
        selected_statuses = self.request.GET.getlist('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        sort_by = self.request.GET.get('sort', '-created_at')
        
        # Filter by status
        if selected_statuses:
            # If statuses are selected, use only those (including Complete if selected)
            status_objects = Status.objects.filter(status__in=selected_statuses)
            queryset = queryset.filter(status__in=status_objects)
        else:
            # Default: exclude complete status when no statuses are selected
            complete_status = Status.objects.filter(status='Complete').first()
            if complete_status:
                queryset = queryset.exclude(status=complete_status)
        
        if date_from:
            queryset = queryset.filter(status_timestamp__gte=date_from)
        if date_to:
            queryset = queryset.filter(status_timestamp__lte=date_to)
        
        # Apply sorting
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['selected_statuses'] = self.request.GET.getlist('status')
        context['date_from'] = self.request.GET.get('date_from')
        context['date_to'] = self.request.GET.get('date_to')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        context['is_current_view'] = True
        return context


class SectioningRequestsCurrentView(ListView):
    model = SectioningRequest
    template_name = 'requests_app/sectioning_requests.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = SectioningRequest.objects.all()
        selected_statuses = self.request.GET.getlist('status')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        sort_by = self.request.GET.get('sort', '-created_at')
        
        # Filter by status
        if selected_statuses:
            # If statuses are selected, use only those (including Complete if selected)
            status_objects = Status.objects.filter(status__in=selected_statuses)
            queryset = queryset.filter(status__in=status_objects)
        else:
            # Default: exclude complete status when no statuses are selected
            complete_status = Status.objects.filter(status='Complete').first()
            if complete_status:
                queryset = queryset.exclude(status=complete_status)
        
        if date_from:
            queryset = queryset.filter(status_timestamp__gte=date_from)
        if date_to:
            queryset = queryset.filter(status_timestamp__lte=date_to)
        
        # Apply sorting
        return queryset.order_by(sort_by)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = Status.objects.exclude(status='submitted').order_by('status')
        context['selected_statuses'] = self.request.GET.getlist('status')
        context['date_from'] = self.request.GET.get('date_from')
        context['date_to'] = self.request.GET.get('date_to')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        context['is_current_view'] = True
        return context


# Request Log Selection Views
class RequestLogSelectionView(TemplateView):
    """View to show all request IDs that have change logs for selection"""
    template_name = 'requests_app/request_log_selection.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get all request IDs that have change logs for each request type
        staining_requests_with_logs = StainingRequestChangeLog.objects.values_list('request_id', flat=True).distinct()
        embedding_requests_with_logs = EmbeddingRequestChangeLog.objects.values_list('request_id', flat=True).distinct()
        sectioning_requests_with_logs = SectioningRequestChangeLog.objects.values_list('request_id', flat=True).distinct()
        
        # Get the actual request objects for display
        context['staining_requests'] = StainingRequest.objects.filter(key__in=staining_requests_with_logs).order_by('-key')
        context['embedding_requests'] = EmbeddingRequest.objects.filter(key__in=embedding_requests_with_logs).order_by('-key')
        context['sectioning_requests'] = SectioningRequest.objects.filter(key__in=sectioning_requests_with_logs).order_by('-key')
        
        return context


# Request History Views
class StainingRequestHistoryView(ListView):
    model = StainingRequestChangeLog
    template_name = 'requests_app/request_history.html'
    context_object_name = 'change_logs'
    paginate_by = 20

    def get_queryset(self):
        request_id = self.kwargs['pk']
        return StainingRequestChangeLog.get_request_history(request_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request_id = self.kwargs['pk']
        try:
            context['request'] = Request.objects.get(pk=request_id)
            context['request_type'] = 'staining'
        except Request.DoesNotExist:
            context['request'] = None
        return context


class EmbeddingRequestHistoryView(ListView):
    model = EmbeddingRequestChangeLog
    template_name = 'requests_app/request_history.html'
    context_object_name = 'change_logs'
    paginate_by = 20

    def get_queryset(self):
        request_id = self.kwargs['pk']
        return EmbeddingRequestChangeLog.get_request_history(request_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request_id = self.kwargs['pk']
        try:
            context['request'] = EmbeddingRequest.objects.get(pk=request_id)
            context['request_type'] = 'embedding'
        except EmbeddingRequest.DoesNotExist:
            context['request'] = None
        return context


class SectioningRequestHistoryView(ListView):
    model = SectioningRequestChangeLog
    template_name = 'requests_app/request_history.html'
    context_object_name = 'change_logs'
    paginate_by = 20

    def get_queryset(self):
        request_id = self.kwargs['pk']
        return SectioningRequestChangeLog.get_request_history(request_id)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request_id = self.kwargs['pk']
        try:
            context['request'] = SectioningRequest.objects.get(pk=request_id)
            context['request_type'] = 'sectioning'
        except SectioningRequest.DoesNotExist:
            context['request'] = None
        return context

